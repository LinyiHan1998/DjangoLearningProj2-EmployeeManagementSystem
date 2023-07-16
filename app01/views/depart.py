from django.shortcuts import render,HttpResponse,redirect
from openpyxl import load_workbook


from app01.models import Department,UserInfo,PrettyNum

from django import forms

from app01.utils.pagination import Pagination
def depart_list(request):

    data_list = Department.objects.all()
    page_obj = Pagination(request,data_list,'')
    page_string = page_obj.html()
    page_query = page_obj.queryset

    return render(request, 'depart_list.html',{"data_list":page_query,"page_string":page_string})
def depart_add(request):
    if request.method == 'GET':
        return render(request,'change.html',{"title":"Add Department"})
    newDep = request.POST.get("Title")
    Department.objects.create(title = newDep)
    return redirect("/depart/list")

def depart_delete(request):
    nid = request.GET.get('nid')
    Department.objects.filter(id = nid).delete()
    return redirect("/depart/list")

def depart_edit(request,nid):
    if request.method =='GET':
        title = Department.objects.filter(id=nid).first()
        print(title.id,title.title)
        return render(request,'depart_edit.html',{"item": title})
    new = request.POST.get('Title')
    Department.objects.filter(id=nid).update(title = new)
    return redirect("/depart/list")

def depart_multi(request):
    #获取用户上传的文件对象
    file_obj = request.FILES.get("exc")
    #对象传递给openpyxl，由openpyxl来读取文件内容
    wb = load_workbook(file_obj)
    #取文件的每个sheet
    sheet = wb.worksheets[0]
    #循环获取每一行数据，从第二行开始
    for row in sheet.iter_rows(min_row=2):
        if Department.objects.filter(title = row[0].value[0:4]).exists():
            continue
        Department.objects.create(title = row[0].value[0:4])
    return redirect('/depart/list')
